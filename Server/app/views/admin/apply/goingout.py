from openpyxl import Workbook

from flask import Response, send_from_directory
from flask_jwt_extended import get_jwt_identity, jwt_required
from flask_restful import Resource
from flasgger import swag_from

from app.docs.admin.apply.goingout import *
from app.models.account import AdminModel, StudentModel
from utils.apply_excel_manager import get_cells


class GoingoutDownload(Resource):
    @swag_from(GOINGOUT_DOWNLOAD_GET)
    @jwt_required
    def get(self):
        """
        외출신청 엑셀 다운로드
        """
        admin = AdminModel.objects(id=get_jwt_identity()).first()
        if not admin:
            return Response('', 403)

        wb = Workbook()
        ws = wb.active

        ws['B2'] = ws['F2'] = ws['J2'] = ws['N2'] = ws['B25'] = ws['F25'] = ws['J25'] = ws['N25'] = ws['B47'] = ws['F47'] = ws['J47'] = ws['N47'] = '학번'
        ws['C2'] = ws['G2'] = ws['K2'] = ws['O2'] = ws['C25'] = ws['G25'] = ws['K25'] = ws['O25'] = ws['C47'] = ws['G47'] = ws['K47'] = ws['O47'] = '이름'

        for student in StudentModel.objects:
            number_cell, name_cell, status_cell = get_cells(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name

            goingout_apply = student.goingout_apply

            if goingout_apply.on_saturday and goingout_apply.on_sunday:
                status = '토요일, 일요일 외출'
            elif goingout_apply.on_saturday:
                status = '토요일 외출'
            elif goingout_apply.on_sunday:
                status = '일요일 외출'
            else:
                status = ''

            ws[status_cell] = status

        wb.save('goingout.xlsx')
        wb.close()

        return send_from_directory('../', 'goingout.xlsx')
